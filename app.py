
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
import pandas as pd
import os
import io
import re
import difflib
import collections
from optimizer import run_optimization_from_db

def fuzzy_match_course(target_code, courses):
    if not target_code: return None
    target_norm = re.sub(r'[\s\-]+', '', target_code).upper()
    
    # Exact normal match
    for c in courses:
        if c.code and re.sub(r'[\s\-]+', '', c.code).upper() == target_norm:
            return c
            
    # Fuzzy match using difflib
    codes_map = {re.sub(r'[\s\-]+', '', c.code).upper(): c for c in courses if c.code}
    matches = difflib.get_close_matches(target_norm, codes_map.keys(), n=1, cutoff=0.85)
    if matches:
        return codes_map[matches[0]]
        
    return None

app = Flask(__name__)
app.config['SECRET_KEY'] = 'apo-secret-key-123'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///apo_data.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'connect_args': {
        'timeout': 30  # SQLite busy_timeout in seconds
    }
}

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

db = SQLAlchemy(app)

# --- Database Models ---

# Association Table for Course <-> Faculty (Many-to-Many)
course_faculty = db.Table('course_faculty',
    db.Column('course_id', db.Integer, db.ForeignKey('course.id'), primary_key=True),
    db.Column('faculty_id', db.Integer, db.ForeignKey('faculty.id'), primary_key=True)
)

class Faculty(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    gender = db.Column(db.String(10), nullable=True)

class Course(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=True) # Can be null if not found
    name = db.Column(db.String(100), nullable=False)
    
    # M2M Relationship
    faculties = db.relationship('Faculty', secondary=course_faculty, lazy='subquery',
        backref=db.backref('courses', lazy=True))
        
    is_frozen = db.Column(db.Boolean, default=False)

class Student(db.Model):
    id = db.Column(db.String(50), primary_key=True) 
    name = db.Column(db.String(100))

class StudentChoice(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.String(50), db.ForeignKey('student.id'), nullable=False)
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'), nullable=False)
    course = db.relationship('Course')

class ScheduleResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.String(50), db.ForeignKey('student.id'))
    course_code = db.Column(db.String(20))
    section = db.Column(db.String(5)) # 'A' or 'B'
    student = db.relationship('Student') 

class TimetableResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    course_code = db.Column(db.String(20))
    course_name = db.Column(db.String(100))
    section = db.Column(db.String(5))
    day = db.Column(db.Integer)  # 0 to 6
    slot = db.Column(db.Integer) # 0 to 6
    faculty_names = db.Column(db.String(200)) # Stored as comma separated just for display

# --- Routes ---

@app.route('/')
def dashboard():
    s_count = Student.query.count()
    c_count = Course.query.count()
    f_count = Faculty.query.count()
    res_count = ScheduleResult.query.count()
    return render_template('dashboard.html', s_count=s_count, c_count=c_count, f_count=f_count, res_count=res_count)

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    cons_path = os.path.join(app.config['UPLOAD_FOLDER'], 'consolidated.xlsx')
    fac_path = os.path.join(app.config['UPLOAD_FOLDER'], 'faculty.xlsx')

    if request.method == 'POST':
        cons_file = request.files.get('consolidated_file')
        fac_file = request.files.get('faculty_file')
        
        uploaded_any = False
        if cons_file and cons_file.filename:
            cons_file.save(cons_path)
            uploaded_any = True
        if fac_file and fac_file.filename:
            fac_file.save(fac_path)
            uploaded_any = True
            
        if not uploaded_any:
            flash("No file was selected for upload.", "warning")
            return redirect(url_for('upload'))
            
        if os.path.exists(cons_path) and os.path.exists(fac_path):
            # RESET DATABASE
            db.drop_all()
            db.create_all()
            
            try:
                # 1. PROCESS STUDENT CHOICES (Consolidated)
                xl = pd.ExcelFile(cons_path)
                IGNORE_SHEETS = {"Sheet1", "Short_Description", "Consolidated"}
                
                count_sheets = 0
                for sheet in xl.sheet_names:
                    if sheet in IGNORE_SHEETS: continue
                    
                    df_sheet = pd.read_excel(xl, sheet_name=sheet)
                    df_sheet.columns = df_sheet.columns.str.strip()
                    
                    id_col_found = None
                    for c in df_sheet.columns:
                        if c.lower() in ("student id", "student ids"):
                            id_col_found = c
                            break
                    
                    # Also try finding the name column dynamically
                    name_col_found = None
                    for c in df_sheet.columns:
                        if c.lower() in ("student name", "student names"):
                            name_col_found = c
                            break

                    if not id_col_found: continue
                        
                    sheet_clean = sheet.strip()
                    course = Course.query.filter_by(code=sheet_clean).first()
                    if not course:
                        course = Course(code=sheet_clean, name=sheet_clean) 
                        db.session.add(course)
                        db.session.commit()
                    
                    count_sheets += 1
                    
                    for _, row in df_sheet.iterrows():
                        sid = str(row[id_col_found]).strip()
                        sname = str(row[name_col_found]).strip() if name_col_found and name_col_found in df_sheet.columns else ""
                        if sname.lower() == 'nan': sname = ""
                        
                        student = Student.query.get(sid)
                        if not student:
                            student = Student(id=sid, name=sname)
                            db.session.add(student)
                        elif not student.name and sname:
                            student.name = sname
                        
                        choice = StudentChoice(student_id=sid, course_id=course.id)
                        db.session.add(choice)
                
                db.session.commit()
                flash(f"Student Data Imported from {count_sheets} courses.", "success")

                # 2. PROCESS FACULTY & COURSE DETAILS
                xl_fac = pd.ExcelFile(fac_path)
                
                # A. Course Details (Exam Sheet) - Updates Titles
                if 'Exam' in xl_fac.sheet_names:
                    df_exam = pd.read_excel(xl_fac, 'Exam')
                    cols = [str(c).strip() for c in df_exam.columns]
                    c_idx, t_idx = -1, -1
                    for i, c in enumerate(cols):
                        if 'Course' in c and 'No' in c: c_idx = i
                        if 'Title' in c: t_idx = i
                        
                    if c_idx != -1 and t_idx != -1:
                        for _, row in df_exam.iterrows():
                            code = str(row[c_idx]).strip()
                            title = str(row[t_idx]).strip()
                            if code != "nan" and title != "nan":
                                crs = Course.query.filter_by(code=code).first()
                                if not crs:
                                    crs = fuzzy_match_course(code, Course.query.all())
                                if crs: crs.name = title
                                else:
                                    crs_by_name = Course.query.filter_by(name=code).first() 
                                    if crs_by_name: crs_by_name.name = title
                        db.session.commit()
                
                # B. Faculty List (Legacy Sheet)
                if 'Faculty List' in xl_fac.sheet_names:
                    df_fac = pd.read_excel(xl_fac, 'Faculty List')
                    valid_cols = df_fac.columns[1:] 
                    for col_name in valid_cols:
                        if "Unnamed" in str(col_name): continue 
                        header_key = str(col_name).strip()
                        if not header_key: continue
                        
                        matched_course = fuzzy_match_course(header_key, Course.query.all())
                        if matched_course:
                            candidates = [matched_course]
                        else:
                            norm_header = re.sub(r'[\s\-]+', '', header_key).upper()
                            candidates = []
                            for c in Course.query.all():
                                if (c.code and norm_header in re.sub(r'[\s\-]+', '', c.code).upper()) or \
                                   (c.name and norm_header in re.sub(r'[\s\-]+', '', c.name).upper()):
                                    candidates.append(c)
                            
                        for course in candidates:
                            col_values = df_fac[col_name].dropna().unique()
                            for val in col_values:
                                raw_val = str(val).strip()
                                # Split by comma in case "Prof A, Prof B" is in this sheet too
                                fnames = [n.strip() for n in raw_val.split(',')]
                                
                                for fname in fnames:
                                    try: float(fname); continue
                                    except ValueError: pass
                                    if fname[0].isdigit() or len(fname) < 2: continue
                                    if fname.lower() in ['session', 'lecture', 'guest lecture', 'student name', 'nan']: continue
                                    
                                    faculty = Faculty.query.filter_by(name=fname).first()
                                    if not faculty:
                                        faculty = Faculty(name=fname)
                                        db.session.add(faculty)
                                        db.session.commit()
                                    if faculty not in course.faculties:
                                        course.faculties.append(faculty)
                    db.session.commit()

                # C. Course Details (New Sheet - Parsing & Cleaning)
                if 'Course Details' in xl_fac.sheet_names:
                    df_cd = pd.read_excel(xl_fac, 'Course Details', header=1)
                    df_cd.columns = df_cd.columns.str.strip()
                    
                    code_col, fac_col, name_col = None, None, None
                    for c in df_cd.columns:
                        if "Course Code" in c: code_col = c
                        if "Faculty Name" in c: fac_col = c
                        if "Course Name" in c: name_col = c
                        
                    if code_col and fac_col:
                        for _, row in df_cd.iterrows():
                            raw_code = str(row[code_col]).strip()
                            fac_name_str = str(row[fac_col]).strip()
                            if raw_code == 'nan' or fac_name_str == 'nan': continue
                            
                            clean_code = raw_code.split('-')[0].strip()
                            course = Course.query.filter_by(code=clean_code).first()
                            
                            if not course:
                                course = fuzzy_match_course(clean_code, Course.query.all())
                            
                            if not course:
                                course = Course(code=clean_code, name=clean_code)
                                db.session.add(course)
                                db.session.commit()
                                        
                            if course:
                                if name_col and str(row[name_col]) != 'nan':
                                    raw_name = str(row[name_col]).strip()
                                    clean_name = re.sub(r'\s*\([A-Za-z0-9]\)$', '', raw_name)
                                    course.name = clean_name
                                
                                # Strict Split logic
                                fac_names = [n.strip() for n in fac_name_str.split(',')]
                                for fname in fac_names:
                                    if not fname or fname.lower() in ['nan', '']: continue
                                    faculty = Faculty.query.filter_by(name=fname).first()
                                    if not faculty:
                                        faculty = Faculty(name=fname)
                                        db.session.add(faculty)
                                        db.session.commit()
                                    if faculty not in course.faculties:
                                        course.faculties.append(faculty)
                        db.session.commit()
                
                flash("Data Upload, Reset, and Processing Complete!", "success")
                
            except Exception as e:
                flash(f"Error processing files: {str(e)}", "danger")
        else:
            flash("File saved. Please ensure both files are uploaded to initialize the system.", "info")

        return redirect(url_for('upload'))
        
    cons_exists = os.path.exists(cons_path)
    fac_exists = os.path.exists(fac_path)
    return render_template('upload.html', cons_exists=cons_exists, fac_exists=fac_exists)

@app.route('/courses', methods=['GET', 'POST'])
def courses():
    if request.method == 'POST':
        cid = request.form.get('id')
        code = request.form.get('code')
        name = request.form.get('name')
        fids = request.form.getlist('faculty_ids') 
        frozen = request.form.get('is_frozen') == 'on'
        
        course = Course.query.get(cid)
        if course:
            course.code = code
            course.name = name
            course.is_frozen = frozen
            
            course.faculties = [] 
            for fid in fids:
                fac = Faculty.query.get(int(fid))
                if fac:
                    course.faculties.append(fac)
            
            db.session.commit()
            flash("Course Updated", "success")
        return redirect(url_for('courses'))
        
    all_courses = Course.query.all()
    all_faculty = Faculty.query.all()
    return render_template('courses.html', courses=all_courses, faculty=all_faculty)

@app.route('/courses/toggle_freeze', methods=['POST'])
def toggle_freeze():
    course_id = request.form.get('course_id')
    course = Course.query.get(course_id)
    if course:
        course.is_frozen = not course.is_frozen
        db.session.commit()
        return jsonify({'success': True, 'is_frozen': course.is_frozen})
    return jsonify({'success': False}), 404

@app.route('/faculty')
def faculty():
    all_faculty = Faculty.query.all()
    all_courses = Course.query.all()
    return render_template('faculty.html', faculty=all_faculty, courses=all_courses)

@app.route('/faculty/add', methods=['POST'])
def faculty_add():
    name = request.form.get('name')
    gender = request.form.get('gender')
    course_ids = request.form.getlist('course_ids')
    
    if name:
        fac = Faculty(name=name, gender=gender)
        db.session.add(fac)
        
        for cid in course_ids:
            c = Course.query.get(int(cid))
            if c:
                fac.courses.append(c)
                
        db.session.commit()
        flash("Faculty Added", "success")
    return redirect(url_for('faculty'))

@app.route('/faculty/update', methods=['POST'])
def faculty_update():
    fid = request.form.get('id')
    name = request.form.get('name')
    gender = request.form.get('gender')
    if fid and name:
        fac = Faculty.query.get(fid)
        if fac:
            fac.name = name
            fac.gender = gender
            db.session.commit()
            flash("Faculty Updated", "success")
    return redirect(url_for('faculty'))

@app.route('/faculty/update_gender_ajax', methods=['POST'])
def faculty_update_gender_ajax():
    fid = request.form.get('id')
    gender = request.form.get('gender')
    if fid:
        fac = Faculty.query.get(fid)
        if fac:
            fac.gender = gender
            db.session.commit()
            return jsonify({'status': 'success', 'gender': gender})
    return jsonify({'status': 'error'}), 400

@app.route('/faculty/delete', methods=['POST'])
def faculty_delete():
    fid = request.form.get('id')
    if fid:
        fac = Faculty.query.get(fid)
        if fac:
            # Check dependencies? For now, just remove from courses
            # SQLAlchemy M2M should handle the association table if configured, but let's be safe
            # Actually, deleting the faculty object will cascade delete the association in most configs
            # If not, we might error. Let's assume standard behavior or clear manually if needed.
             db.session.delete(fac)
             db.session.commit()
             flash("Faculty Deleted", "success")
    return redirect(url_for('faculty'))

@app.route('/optimize')
def optimize():
    try:
        _run_optimization_logic()
        flash(f"Section Optimization Completed!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Optimization Error: {str(e)}", "danger")
    return redirect(url_for('results'))

def _run_optimization_logic():
    courses = Course.query.all()
    choices = StudentChoice.query.all()
    
    courses_data = [{'id': c.id, 'code': c.code, 'is_frozen': c.is_frozen} for c in courses]
    
    assignments_data = [] 
    for ch in choices:
        key = ch.course.code if ch.course.code else ch.course.name
        assignments_data.append({'student_id': ch.student_id, 'course_code': key})
    
    results, obj_score = run_optimization_from_db(courses_data, assignments_data)
    
    db.session.query(ScheduleResult).delete(synchronize_session=False) 
    for r in results:
        res = ScheduleResult(student_id=r['student_id'], course_code=r['course_code'], section=r['section'])
        db.session.add(res)
    db.session.commit()
    return obj_score

@app.route('/results')
def results():
    results = ScheduleResult.query.limit(100).all() 
    return render_template('results.html', results=results)

from scheduler import generate_master_timetable, assign_students_to_sections

@app.route('/schedule', methods=['GET', 'POST'])
def schedule():
    if request.method == 'POST':
        try:
            # 1. Prepare Data for Optimizer
            courses = Course.query.all()
            choices = StudentChoice.query.all()
            
            courses_data = {c.code: [] for c in courses if c.code}
            student_to_courses = collections.defaultdict(set)
            for ch in choices:
                if ch.course and ch.course.code:
                    courses_data[ch.course.code].append(ch.student_id)
                    student_to_courses[ch.student_id].add(ch.course.code)
            
            frozen_courses = {c.code for c in courses if c.is_frozen}
            
            # 2. Run Section Optimization (Dynamic Anchor Split)
            from optimizer import optimize_splits
            assignments, obj_score = optimize_splits(courses_data, student_to_courses, frozen_courses=frozen_courses)
            
            # 3. Save Section Assignments
            db.session.query(ScheduleResult).delete(synchronize_session=False)
            for c_code, s_map in assignments.items():
                for sid, sec_idx in s_map.items():
                    sec_name = 'A' if sec_idx == 0 else 'B'
                    db.session.add(ScheduleResult(student_id=sid, course_code=c_code, section=sec_name))
            db.session.commit()
            
            # 4. Compute Section Overlap Matrix (for Scheduler)
            import numpy as np
            section_students = collections.defaultdict(set)
            for c_code, s_map in assignments.items():
                for sid, sec_idx in s_map.items():
                    sec_name = 'A' if sec_idx == 0 else 'B'
                    section_students[f"{c_code}-{sec_name}"].add(sid)
            
            all_sec_ids = sorted(list(section_students.keys()))
            n = len(all_sec_ids)
            overlap_matrix = np.zeros((n, n), dtype=int)
            for i in range(n):
                for j in range(i + 1, n):
                    inter = len(section_students[all_sec_ids[i]] & section_students[all_sec_ids[j]])
                    overlap_matrix[i, j] = inter
                    overlap_matrix[j, i] = inter
            
            overlap_df = pd.DataFrame(overlap_matrix, index=all_sec_ids, columns=all_sec_ids)
            
            # 5. Run Backtracking Scheduler
            master_tt = generate_master_timetable(db.session, Course, Faculty, StudentChoice, overlap_matrix=overlap_df)
            
            if master_tt:
                db.session.query(TimetableResult).delete(synchronize_session=False)
                for t in master_tt:
                    db.session.add(TimetableResult(
                        course_code=t['course_code'],
                        course_name=t.get('course_name', ''),
                        section=t['section'],
                        day=t['day'],
                        slot=t['slot'],
                        faculty_names=t['faculty_names']
                    ))
                db.session.commit()
                flash(f"Schedule generated successfully! (Objective: {obj_score})", "success")
            else:
                flash("Scheduler failed to find a valid timetable. Try adjusting constraints.", "danger")
                
        except Exception as e:
            db.session.rollback()
            flash(f"Error during scheduling: {str(e)}", "danger")
            
        return redirect(url_for('timetable_view'))
    
    # Just render a page with a button to trigger scheduling
    return render_template('schedule_trigger.html')

@app.route('/api/swap_slots', methods=['POST'])
def swap_slots():
    data = request.json
    d1, s1 = int(data.get('day1')), int(data.get('slot1'))
    d2, s2 = int(data.get('day2')), int(data.get('slot2'))
    
    # Get all entries in both slots
    items1 = TimetableResult.query.filter_by(day=d1, slot=s1).all()
    items2 = TimetableResult.query.filter_by(day=d2, slot=s2).all()
    
    # Swap them
    for item in items1:
        item.day = d2
        item.slot = s2
    for item in items2:
        item.day = d1
        item.slot = s1
        
    db.session.commit()
    return jsonify({'success': True})

@app.route('/timetable')
def timetable_view():
    results = TimetableResult.query.all()
    
    # Organize by day and slot
    # structure: {day: {slot: [items]}}
    tt = collections.defaultdict(lambda: collections.defaultdict(list))
    for r in results:
        tt[r.day][r.slot].append(r)
        
    return render_template('timetable.html', timetable=tt)

@app.route('/download')
def download():
    import numpy as np
    assignments = ScheduleResult.query.all()
    tt_results = TimetableResult.query.all()
    
    # 1. Prepare Section Data
    data_by_section = {}
    student_sections = {}
    for r in assignments:
        section_name = f"{r.course_code}-{r.section}"
        if section_name not in data_by_section:
            data_by_section[section_name] = []
        s_name = r.student.name if r.student else ""
        data_by_section[section_name].append({'Student ID': r.student_id, 'Student Name': s_name})
        if r.student_id not in student_sections: student_sections[r.student_id] = []
        student_sections[r.student_id].append(section_name)
    
    all_sections = sorted(list(data_by_section.keys()))
    
    # 2. Rescheduling Master (Overlaps)
    n = len(all_sections)
    overlap_matrix = np.zeros((n, n), dtype=int)
    section_index = {sec: i for i, sec in enumerate(all_sections)}
    for sid, secs in student_sections.items():
        for i in range(len(secs)):
            for j in range(i + 1, len(secs)):
                idx1, idx2 = section_index[secs[i]], section_index[secs[j]]
                overlap_matrix[idx1, idx2] += 1
                overlap_matrix[idx2, idx1] += 1
    rescheduling_df = pd.DataFrame(overlap_matrix, index=all_sections, columns=all_sections)
    
    # 3. Master Timetable (Visual Grid view)
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    times = ['09:00-10:30', '10:45-12:15', '12:30-14:00', '15:00-16:30', '16:45-18:15', '18:30-20:00', '20:15-21:45']
    
    # Grid: Times as Rows, Days as Columns
    grid = {time: {day: "" for day in days} for time in times}
    for t in tt_results:
        if 0 <= t.day < len(days) and 0 <= t.slot < len(times):
            d_name = days[t.day]
            s_name = times[t.slot]
            
            # Formatted entry: Code-Section \n Name \n Faculty
            fname = t.faculty_names or ""
            entry = f"{t.course_code}-{t.section}\n{t.course_name}\n({fname})"
            
            if grid[s_name][d_name]:
                grid[s_name][d_name] += f"\n\n{entry}"
            else:
                grid[s_name][d_name] = entry
                
    master_tt_df = pd.DataFrame.from_dict(grid, orient='index')[days]
    
    # 4. Save to Excel with Formatting
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        master_tt_df.to_excel(writer, sheet_name="Master Timetable")
        rescheduling_df.to_excel(writer, sheet_name="Rescheduling Master")
        
        for sec_name in all_sections:
            rows = data_by_section[sec_name]
            df = pd.DataFrame(rows)
            df.sort_values(by=['Student ID'], inplace=True)
            sheet_name = str(sec_name)[:31].replace('/','_').replace(':','_')
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
        # Access openpyxl objects for styling
        workbook = writer.book
        from openpyxl.styles import Alignment, Font, PatternFill
        
        if "Master Timetable" in writer.sheets:
            ws = writer.sheets["Master Timetable"]
            # Bold headers
            for cell in ws[1]: # First row (Days)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            for cell in ws['A']: # First column (Times)
                cell.font = Font(bold=True)
            
            # Wrap text and set width
            for row in ws.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            for col in ['B','C','D','E','F','G','H']:
                ws.column_dimensions[col].width = 30
                
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                     as_attachment=True, download_name="APO_Consolidated_Schedule.xlsx")

import os
if __name__ == '__main__':
    with app.app_context():
        db_dir = os.path.join(app.instance_path)
        if not os.path.exists(db_dir):
            os.makedirs(db_dir)
        db.create_all()
    
    # Use environment-defined port for Render
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=False)

# Refresh Trigger Comment
